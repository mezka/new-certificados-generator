from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook
from time import strftime
from jinja2 import Template
from config import products, defaults
from weasyprint import HTML
from PyPDF2 import PdfFileReader, PdfFileWriter
from os import path, remove, makedirs
from shutil import rmtree
from wand.image import Image as WandImage
import time

def create_temp_directory_if_not_exists():
    try:
        makedirs(defaults['temp_directory_path'])
    except OSError:
        raise
def delete_temp_directory():
    try:
        rmtree(defaults['temp_directory_path'])
    except OSError:
        raise

def load_workbook_rows_into_certificados():
    wb = load_workbook(filename = defaults['excel_workbook_filename'])
    ws = wb.active

    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        keys = row

    certificados = []

    for row in ws.iter_rows(min_row=2, values_only=True):

        certificado = dict(zip(keys, row))

        try:
            certificado['ORDEN_DE_TRABAJO'] = products[certificado['MODELO']]['orden_de_trabajo']
            certificado['CLASIFICACION'] = products[certificado['MODELO']]['clasificacion']
            certificado['FECHA'] = strftime("%d-%m-%Y")
            certificado['_template_filename'] = products[certificado['MODELO']]['template_filename']
            certificado['_pdf_certificate_filename'] = products[certificado['MODELO']]['pdf_certificate_filename']
        except KeyError as e:
            raise

        certificados.append(certificado)

    return certificados

def generate_pdf_from_template(certificado):

    with open(certificado['_template_filename'], 'r') as template_file:
        template_string = template_file.read()

    jinja_template = Template(template_string)

    HTML(string=jinja_template.render(certificado), base_url=path.dirname(certificado['_template_filename'])).write_pdf(defaults['temp_lastpage_pdf_filename'])

def generate_watermark_text(certificado):
    
    with open(defaults['watermark_template_filename'], 'r') as template_file:
        template_string = template_file.read()

    jinja_template = Template(template_string)

    return jinja_template.render(certificado).upper()

def generate_watermark_pdf(watermark_text):

    bg = Image.new("RGBA", (595,842), (255,255,255,0))
    wm = Image.new("RGBA", (800,800), (255, 255, 255, 0))

    draw = ImageDraw.Draw(wm)
    
    draw.multiline_text((0,0), watermark_text, fill=(255, 0, 0, 200), font=ImageFont.truetype(defaults['watermark_font_filename'], 36), align='center')
    
    wm = wm.rotate(45, expand=True, resample=Image.BICUBIC)
    imageBox = wm.getbbox()
    wm = wm.crop(imageBox)

    middle = (int(bg.size[0]/2 - wm.size[0]/2), int(bg.size[1]/2 - wm.size[1]/2))

    bg.paste(wm, middle, wm)
    bg.save(defaults['temp_watermark_png_filename'], "PNG")

    bg.close()
    wm.close()

    with WandImage(filename = defaults['temp_watermark_png_filename']) as img:
        img.format = 'pdf'
        try:
            img.save(filename = defaults['temp_watermark_pdf_filename'])
        except:
            raise


    remove(defaults['temp_watermark_png_filename'])

def concatenate_pdf(first_pdf_filename, resulting_pdf_filename):

    pdf_writer = PdfFileWriter()
    
    pdf_watermark_file = open(defaults['temp_watermark_pdf_filename'], 'rb')
    pdf_watermark_reader = PdfFileReader(pdf_watermark_file, strict=False)
    pdf_watermark_page = pdf_watermark_reader.getPage(0) 

    first_pdf_file = open(first_pdf_filename, 'rb')
    first_pdf_reader = PdfFileReader(first_pdf_file, strict=False)

    for page_num in range(first_pdf_reader.numPages):
        page_obj = first_pdf_reader.getPage(page_num)
        page_obj.mergePage(pdf_watermark_page)
        pdf_writer.addPage(page_obj)

    second_pdf_file = open(defaults['temp_lastpage_pdf_filename'], 'rb')
    second_pdf_reader = PdfFileReader(second_pdf_file)


    for page_num in range(second_pdf_reader.numPages):
        page_obj = second_pdf_reader.getPage(page_num)
        pdf_writer.addPage(page_obj)

    with open(resulting_pdf_filename, 'wb') as dest_pdf_file:
        pdf_writer.write(dest_pdf_file)

    first_pdf_file.close()
    second_pdf_file.close()
    pdf_watermark_file.close()
    
    remove(defaults['temp_lastpage_pdf_filename'])
    remove(defaults['temp_watermark_pdf_filename'])

if __name__ == "__main__":
    
    try:
        certificados = load_workbook_rows_into_certificados()


        for certificado in certificados:
            create_temp_directory_if_not_exists()
            try:
                generate_pdf_from_template(certificado)
                generate_watermark_pdf(generate_watermark_text(certificado))
                concatenate_pdf(certificado['_pdf_certificate_filename'], path.join(defaults['out_directory_path'], f"{certificado['CLIENTE'].replace(' ', '_')}_{certificado['MODELO']}_{certificado['MEDIDA']}.pdf"))
            except Exception as e:
                raise
            finally:
                delete_temp_directory()
    except OSError:
        print('ERROR: Hubo un problema al crear o eliminar el directorio temporal, a continuacion se imprimen los detalles del error:\n')
        raise
    except KeyError:
        print('ERROR: Hubo un problema al leer el codigo de modelo especificado en una de las filas, a continuacion se imprimen los detalles del error:\n')
        raise
